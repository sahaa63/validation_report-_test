import streamlit as st
import pandas as pd
import io
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import base64  # For base64 image encoding

# Check for openpyxl availability
try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    st.error("The 'openpyxl' library is not installed. Please ensure it's included in your requirements.txt and the environment is set up correctly.")
    st.stop()

# Custom CSS for styling with improved contrast
st.markdown("""
    <style>
    .title {
        font-size: 36px;
        color: #FF4B4B;
        text-align: center;
        font-weight: bold;
        margin-bottom: 20px;
    }
    .instructions {
        background-color: #F0F8FF;
        color: #333333;
        padding: 15px;
        border-radius: 10px;
        border-left: 5px solid #4682B4;
        margin-bottom: 20px;
    }
    .file-list {
        background-color: #F5F5F5;
        color: #333333;
        padding: 10px;
        border-radius: 5px;
        margin-top: 10px;
    }
    .stButton>button {
        background-color: #4CAF50;
        color: white;
        border: none;
        padding: 10px 20px;
        border-radius: 5px;
        font-weight: bold;
    }
    .stButton>button:hover {
        background-color: #45A049;
    }
    .success-box {
        background-color: #E6FFE6;
        color: #333333;
        padding: 15px;
        border-radius: 10px;
        border-left: 5px solid #2ECC71;
        margin-top: 20px;
    }
    .error-box {
        background-color: #FFE6E6;
        color: #333333;
        padding: 15px;
        border-radius: 10px;
        border-left: 5px solid #FF4B4B;
        margin-top: 20px;
    }
    </style>
""", unsafe_allow_html=True)

def apply_conditional_formatting(ws, sheet_name, wb):
    temp_buffer = io.BytesIO()
    wb.save(temp_buffer)
    temp_buffer.seek(0)
    df = pd.read_excel(temp_buffer, sheet_name=sheet_name)

    dark_green_fill = PatternFill(start_color='19D119', end_color='19D119', fill_type='solid')
    yellow_fill = PatternFill(start_color='E4E81B', end_color='E4E81B', fill_type='solid')
    dark_red_fill = PatternFill(start_color='E82D1C', end_color='E82D1C', fill_type='solid')

    diff_cols = [col for col in df.columns if col.endswith('_Diff')]
    presence_col_idx = df.columns.get_loc('presence') + 1 if 'presence' in df.columns else None
    
    for col_idx, col_name in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_idx)
        
        if col_name.endswith('_Diff'):
            header_cell = ws[f'{col_letter}1']
            header_cell.number_format = '0.00%'
            
            for row_idx, value in enumerate(df[col_name], 2):
                cell = ws[f'{col_letter}{row_idx}']
                if pd.notna(value):
                    cell.value = value
                    cell.number_format = '0.00%'
                    if value <= 0.1:
                        cell.fill = dark_green_fill
                    elif value <= 0.5:
                        ratio = (value - 0.1) / 0.5
                        r = int(255 + (139 - 255) * ratio)
                        g = int(255 - (255 - 0) * ratio)
                        b = int(0)
                        color = f'{r:02X}{g:02X}{b:02X}'
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    else:
                        cell.fill = dark_red_fill
        
        elif presence_col_idx and col_idx == presence_col_idx:
            for row_idx, value in enumerate(df[col_name], 2):
                cell = ws[f'{col_letter}{row_idx}']
                if value == 'Present in Both':
                    cell.fill = dark_green_fill
                elif value in ['Present in excel', 'Present in PBI']:
                    cell.fill = dark_red_fill

def combine_excel_files(file_list):
    if not file_list or len(file_list) > 10:
        return None, None

    first_filename = os.path.splitext(file_list[0].name)[0]
    base_name = first_filename.split('_')[0]
    output_filename = f"{base_name}_validation_report.xlsx"

    output_buffer = io.BytesIO()
    output_wb = Workbook()
    sheet_order = []
    sheet_name_count = {}

    for uploaded_file in file_list:
        file_bytes = uploaded_file.read()
        try:
            wb = load_workbook(filename=io.BytesIO(file_bytes))
        except Exception as e:
            st.error(f"Error reading file {uploaded_file.name}: {str(e)}")
            return None, None

        for sheet_name in wb.sheetnames:
            base_sheet_name = sheet_name
            if sheet_name in sheet_name_count:
                sheet_name_count[sheet_name] += 1
                new_sheet_name = f"{base_sheet_name}_{sheet_name_count[sheet_name]}"
            else:
                sheet_name_count[sheet_name] = 0
                new_sheet_name = sheet_name

            ws_source = wb[base_sheet_name]
            ws_target = output_wb.create_sheet(title=new_sheet_name)
            for row in ws_source.rows:
                for cell in row:
                    ws_target[cell.coordinate].value = cell.value
            sheet_order.append(new_sheet_name)

    if 'Sheet' in output_wb.sheetnames:
        output_wb.remove(output_wb['Sheet'])
    output_wb._sheets = [output_wb[sheet] for sheet in sheet_order]

    for sheet_name in output_wb.sheetnames:
        apply_conditional_formatting(output_wb[sheet_name], sheet_name, output_wb)

    output_wb.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer, output_filename

# Function to encode local image as base64
def get_base64_image(image_path):
    with open(image_path, "rb") as img_file:
        return base64.b64encode(img_file.read()).decode()

def main():
    st.markdown('<div class="title">Excel File Merger</div>', unsafe_allow_html=True)

    st.markdown("""
    <div class="instructions">
    <h3 style="color: #4682B4;">How to Use:</h3>
    <ul>
        <li>Upload up to 10 Excel files using the button below.</li>
        <li>All sheets from each file will be merged into one output file <strong>in the order you upload them</strong>.</li>
        <li>Duplicate sheet names will get a numeric suffix (e.g., 'Sheet_1').</li>
        <li>The output file will be named using the first file's prefix before the first underscore (e.g., 'Retailer Redemption_validation_report.xlsx').</li>
    </ul>
    </div>
    """, unsafe_allow_html=True)

    uploaded_files = st.file_uploader(
        "Drop Your Excel Files Here!",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
        help="Upload up to 10 Excel files to merge into one. Sheets will appear in upload order.",
        key="file_uploader"
    )

    if uploaded_files:
        if len(uploaded_files) > 10:
            st.markdown(
                '<div class="error-box">Whoops! Maximum 10 files allowed. Please upload fewer files.</div>',
                unsafe_allow_html=True
            )
        else:
            st.markdown(f'<div class="file-list"><strong>Uploaded {len(uploaded_files)} File(s):</strong>', unsafe_allow_html=True)
            for file in uploaded_files:
                st.markdown(f"- {file.name}", unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)

            with st.spinner("Merging your files... Hang tight!"):
                result = combine_excel_files(uploaded_files)
                if result:
                    output_buffer, output_filename = result
                    st.markdown(
                        f'<div class="success-box">Success! Your merged file is ready: <strong>{output_filename}</strong></div>',
                        unsafe_allow_html=True
                    )
                    st.download_button(
                        label="Download Your Merged Excel!",
                        data=output_buffer,
                        file_name=output_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_button"
                    )

    # Fancy Footer with Local Image (Sigmoid_Logo.jpg) in Left Upper Corner
    try:
        image_base64 = get_base64_image("Sigmoid_Logo.jpg")
        image_src = f"data:image/jpeg;base64,{image_base64}"
    except FileNotFoundError:
        # Fallback to placeholder if image not found
        image_src = "https://via.placeholder.com/100"
        st.warning("Sigmoid_Logo.jpg not found in the directory. Using placeholder image.")

    footer_html = f"""
    <div style='background-color: #FFFFFF; color: #000000; padding: 20px; border-radius: 10px; box-shadow: 0 4px 8px rgba(0,0,0,0.2); margin-top: 30px; position: relative;'>
        <img src="{image_src}" alt="Sigmoid Logo" style='position: absolute; top: 10px; left: 10px; width: 100px; height: auto; border-radius: 5px;'>
        <div style='margin-left: 120px;'> <!-- Adjust margin to avoid overlap with logo -->
            <p style='font-size: 16px; font-weight: bold; margin: 10px 0 5px 0;'>Contact Us</p>
            <p style='font-size: 14px; margin: 0;'>
                Email: <a href='mailto:arkaprova@sigmoidanalytics.com' style='color: #1E90FF; text-decoration: none;'>arkaprova@sigmoidanalytics.com</a><br>
                Phone: <span style='color: #FFD700;'>+91 9330492917</span><br>
                Website: <a href='https://merge02.streamlit.app' style='color: #1E90FF; text-decoration: none;'>merge02.streamlit.app</a>
            </p>
        </div>
    </div>
    """
    st.markdown(footer_html, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
