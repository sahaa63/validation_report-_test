import streamlit as st
import pandas as pd
import io
import numpy as np
import os
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# Define the checklist data as a DataFrame
checklist_data = {
    "S.No": range(1, 8),
    "Checklist": [
        "All the columns of excel replicated in PBI (No extra columns)",
        "All the filters of excel replicated in PBI",
        "Filters working as expected (single/multi select as usual)",
        "Column names matching with excel",
        "Currency symbols to be replicated",
        "Pre-applied filters while generating validation report?",
        "Sorting is replicated"
    ],
}
checklist_df = pd.DataFrame(checklist_data)

def generate_validation_report(excel_df, pbi_df):
    # Identify dimensions and measures
    dims = [col for col in excel_df.columns if col in pbi_df.columns and 
            (excel_df[col].dtype == 'object' or '_id' in col.lower() or '_key' in col.lower() or
             '_ID' in col or '_KEY' in col)]

    # Replace empty/blanks in dim columns with 'NAN' string
    excel_df[dims] = excel_df[dims].fillna('NAN')
    pbi_df[dims] = pbi_df[dims].fillna('NAN')

    # Only numeric columns that can be summed are considered as measures
    excel_measures = [col for col in excel_df.columns if col not in dims and np.issubdtype(excel_df[col].dtype, np.number)]
    pbi_measures = [col for col in pbi_df.columns if col not in dims and np.issubdtype(pbi_df[col].dtype, np.number)]
    
    all_measures = list(set(excel_measures) & set(pbi_measures))  # Only measures present in both

    # Aggregate by dimensions to handle duplicate dimensional combinations
    excel_agg = excel_df.groupby(dims)[all_measures].sum().reset_index()
    pbi_agg = pbi_df.groupby(dims)[all_measures].sum().reset_index()

    # Create a unique key by concatenating all dimensions
    excel_agg['unique_key'] = excel_agg[dims].astype(str).agg('-'.join, axis=1).str.upper()
    pbi_agg['unique_key'] = pbi_agg[dims].astype(str).agg('-'.join, axis=1).str.upper()

    # Move 'unique_key' to the first column
    excel_agg = excel_agg[['unique_key'] + [col for col in excel_agg.columns if col != 'unique_key']]
    pbi_agg = pbi_agg[['unique_key'] + [col for col in pbi_agg.columns if col != 'unique_key']]

    # Create the validation report dataframe
    validation_report = pd.DataFrame({'unique_key': list(set(excel_agg['unique_key']) | set(pbi_agg['unique_key']))})

    # Add dimensions
    for dim in dims:
        validation_report[dim] = validation_report['unique_key'].map(dict(zip(excel_agg['unique_key'], excel_agg[dim])))
        validation_report[dim].fillna(validation_report['unique_key'].map(dict(zip(pbi_agg['unique_key'], pbi_agg[dim]))), inplace=True)

    # Determine presence in sheets
    validation_report['presence'] = validation_report['unique_key'].apply(
        lambda key: 'Present in Both' if key in excel_agg['unique_key'].values and key in pbi_agg['unique_key'].values
        else ('Present in excel' if key in excel_agg['unique_key'].values
              else 'Present in PBI')
    )

    # Add measures and calculate differences
    for measure in all_measures:
        validation_report[f'{measure}_excel'] = validation_report['unique_key'].map(dict(zip(excel_agg['unique_key'], excel_agg[measure])))
        validation_report[f'{measure}_PBI'] = validation_report['unique_key'].map(dict(zip(pbi_agg['unique_key'], pbi_agg[measure])))
        
        # Calculate difference (keeping as decimal between 0 and 1)
        validation_report[f'{measure}_Diff'] = np.where(
            (validation_report[f'{measure}_PBI'].fillna(0) == 0) | (validation_report[f'{measure}_excel'].fillna(0) == 0),
            np.where(
                (validation_report[f'{measure}_PBI'].fillna(0) == 0) & (validation_report[f'{measure}_excel'].fillna(0) == 0),
                0,
                1
            ),
            abs(round((validation_report[f'{measure}_PBI'].fillna(0) - validation_report[f'{measure}_excel'].fillna(0)) / 
                     validation_report[f'{measure}_excel'].fillna(0), 4))
        )

    # Reorder columns
    column_order = ['unique_key'] + dims + ['presence'] + \
                   [col for measure in all_measures for col in 
                    [f'{measure}_excel', f'{measure}_PBI', f'{measure}_Diff']]
    validation_report = validation_report[column_order]

    return validation_report, excel_agg, pbi_agg

def column_checklist(excel_df, pbi_df):
    excel_columns = excel_df.columns.tolist()
    pbi_columns = pbi_df.columns.tolist()

    checklist_df = pd.DataFrame({
        'excel Columns': excel_columns + [''] * (max(len(pbi_columns), len(excel_columns)) - len(excel_columns)),
        'PowerBI Columns': pbi_columns + [''] * (max(len(pbi_columns), len(excel_columns)) - len(pbi_columns))
    })

    checklist_df['Match'] = checklist_df.apply(lambda row: row['excel Columns'] == row['PowerBI Columns'], axis=1)
    
    return checklist_df

def generate_diff_checker(validation_report):
    diff_columns = [col for col in validation_report.columns if col.endswith('_Diff')]

    diff_checker = pd.DataFrame({
        'Diff Column Name': diff_columns,
        'Percentage Difference': [f"{validation_report[col].mean()*100:.2f}%" for col in diff_columns]
    })

    presence_summary = {
        'Diff Column Name': 'All rows present in both',
        'Percentage Difference': 'Yes' if all(validation_report['presence'] == 'Present in Both') else 'No'
    }
    diff_checker = pd.concat([diff_checker, pd.DataFrame([presence_summary])], ignore_index=True)

    return diff_checker

def apply_conditional_formatting(ws, validation_report):
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

    diff_cols = [col for col in validation_report.columns if col.endswith('_Diff')]
    
    for col_idx, col_name in enumerate(validation_report.columns, 1):
        if col_name.endswith('_Diff'):
            col_letter = get_column_letter(col_idx)
            
            # Format header
            header_cell = ws[f'{col_letter}1']
            header_cell.number_format = '0.00%'
            
            # Apply formatting to data cells
            for row_idx, value in enumerate(validation_report[col_name], 2):
                cell = ws[f'{col_letter}{row_idx}']
                if pd.notna(value):
                    cell.value = value
                    cell.number_format = '0.00%'
                    
                    if value <= 0.25:
                        cell.fill = green_fill
                    elif value <= 0.75:
                        ratio = (value - 0.25) / 0.5
                        r = int(255 + (255 - 255) * ratio)
                        g = int(255 - (255 - 165) * ratio)
                        b = int(0 + (0 - 0) * ratio)
                        color = f'{r:02X}{g:02X}{b:02X}'
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    else:
                        cell.fill = orange_fill

def main():
    st.title("Validation Report Generator")

    st.markdown("""
    **Important Assumptions:**
    1. Upload the Excel file with two sheets: "excel" and "PBI".
    2. Make sure the column names are similar in both sheets.
    3. If there are ID/Key/Code columns, make sure the ID or Key columns contains "_ID" or "_KEY" (case insensitive)
    """)

    uploaded_file = st.file_uploader("Upload Excel file", type="xlsx")

    if uploaded_file is not None:
        try:
            xls = pd.ExcelFile(uploaded_file)
            excel_df = pd.read_excel(xls, 'excel')
            pbi_df = pd.read_excel(xls, 'PBI')

            # Convert all string columns to uppercase and trim whitespace
            excel_df = excel_df.apply(lambda x: x.str.upper().str.strip() if x.dtype == "object" else x)
            pbi_df = pbi_df.apply(lambda x: x.str.upper().str.strip() if x.dtype == "object" else x)

            validation_report, excel_agg, pbi_agg = generate_validation_report(excel_df, pbi_df)
            column_checklist_df = column_checklist(excel_df, pbi_df)
            diff_checker_df = generate_diff_checker(validation_report)

            st.subheader("Validation Report Preview")
            # Convert diff columns to percentage for Streamlit display
            display_report = validation_report.copy()
            for col in display_report.columns:
                if col.endswith('_Diff'):
                    display_report[col] = display_report[col].apply(lambda x: f"{x*100:.2f}%" if pd.notna(x) else x)
            st.dataframe(display_report)

            # Generate Excel file for download
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                validation_report.to_excel(writer, sheet_name='validation_report', index=False)
                ws = writer.sheets['validation_report']
                apply_conditional_formatting(ws, validation_report)

            output.seek(0)
            original_filename = os.path.splitext(uploaded_file.name)[0]
            new_file_name = f"{original_filename}_validation_report.xlsx"
            st.download_button(
                label="Download Excel Report",
                data=output,
                file_name=new_file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            st.markdown('---')
            st.markdown('--')
        except Exception as e:
            st.error(f"An error occurred: {str(e)}")

if __name__ == "__main__":
    main()
